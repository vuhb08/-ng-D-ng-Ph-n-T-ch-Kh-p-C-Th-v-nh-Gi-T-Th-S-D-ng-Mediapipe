import cv2
import mediapipe as mp
import numpy as np
import tkinter as tk
from tkinter import ttk, simpledialog
from PIL import Image, ImageTk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from matplotlib import rcParams
import openpyxl
from datetime import datetime
from tkinter import messagebox
import os
import glob
from scipy.interpolate import splprep, splev
from collections import deque

# =========================
# Biến lưu thông tin bệnh nhân
# =========================
patient_info = {
    "name": "",
    "dob": "",
    "joint": "",
    "condition": ""
}

# Biến toàn cục để lưu metrics và góc
latest_metrics = None
latest_angle_deg = None

# =========================
# Mediapipe setup
# =========================
mp_drawing = mp.solutions.drawing_utils
mp_pose = mp.solutions.pose
pose = mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.7)

# =========================
# GUI setup
# =========================
root = tk.Tk()
root.title("Theo dõi Khớp Cơ thể - Điều khiển")
plt.style.use('dark_background')
rcParams['axes.edgecolor'] = 'white'
rcParams['text.color'] = 'white'
rcParams['axes.labelcolor'] = 'white'
rcParams['xtick.color'] = 'white'
rcParams['ytick.color'] = 'white'

# Tạo cửa sổ riêng cho video
video_window = tk.Toplevel(root)
video_window.title("Video Theo dõi")
video_window.configure(bg='#2e2e2e')

# Tạo cửa sổ riêng cho So sánh thực tế
compare_window = tk.Toplevel(root)
compare_window.title("So sánh thực tế")
compare_window.configure(bg='#2e2e2e')

# =========================
# Giao diện nhập thông tin bệnh nhân
# =========================
info_frame = tk.Frame(root, bg='#2e2e2e')
info_frame.pack(padx=10, pady=10, fill=tk.X)

# Thêm Combobox lịch sử bệnh nhân
tk.Label(info_frame, text="Lịch sử bệnh nhân:", fg='white', bg='#2e2e2e').grid(row=0, column=0, sticky='w')
combo_patient_history = ttk.Combobox(info_frame, state="readonly")
combo_patient_history.grid(row=0, column=1, padx=5, pady=2)
combo_patient_history.set("Chọn bệnh nhân...")

# Tải danh sách bệnh nhân từ các file Excel
def load_patient_history():
    patient_files = glob.glob("*.xlsx")
    patient_names = []
    for file in patient_files:
        try:
            wb = openpyxl.load_workbook(file)
            if "Thông tin bệnh nhân" in wb.sheetnames:
                info_sheet = wb["Thông tin bệnh nhân"]
                name = info_sheet["B1"].value
                if name:
                    patient_names.append(name)
        except Exception:
            continue
    combo_patient_history['values'] = patient_names
    if patient_names:
        combo_patient_history.set("Chọn bệnh nhân...")
    else:
        combo_patient_history.set("Không có bệnh nhân")

# Gọi hàm tải lịch sử bệnh nhân khi khởi động
load_patient_history()

tk.Label(info_frame, text="Tên bệnh nhân:", fg='white', bg='#2e2e2e').grid(row=1, column=0, sticky='w')
entry_name = tk.Entry(info_frame, width=30)
entry_name.grid(row=1, column=1, padx=5, pady=2)

tk.Label(info_frame, text="Ngày sinh (dd/mm/yyyy):", fg='white', bg='#2e2e2e').grid(row=2, column=0, sticky='w')
entry_dob = tk.Entry(info_frame, width=30)
entry_dob.grid(row=2, column=1, padx=5, pady=2)

tk.Label(info_frame, text="Khớp kiểm tra:", fg='white', bg='#2e2e2e').grid(row=3, column=0, sticky='w')
entry_joint = tk.Entry(info_frame, width=30)
entry_joint.grid(row=3, column=1, padx=5, pady=2)

tk.Label(info_frame, text="Tình trạng:", fg='white', bg='#2e2e2e').grid(row=4, column=0, sticky='w')
entry_condition = tk.Entry(info_frame, width=30)
entry_condition.grid(row=4, column=1, padx=5, pady=2)

# =========================
# Thêm tính năng thước đo
# =========================
measuring = False
start_x, start_y = None, None
end_x, end_y = None, None
original_image = None
orig_w, orig_h = None, None
label_w, label_h = None, None
measure_message_shown = False

def toggle_measure_mode():
    global measuring, start_x, start_y, end_x, end_y, measure_message_shown
    measuring = not measuring
    if measuring:
        if not measure_message_shown:
            messagebox.showinfo("Hướng dẫn", "Kéo chuột trên video để đo khoảng cách.")
            measure_message_shown = True
    else:
        start_x, start_y, end_x, end_y = None, None, None, None
        measure_message_shown = False

def start_measure(event):
    global start_x, start_y, measuring
    if measuring:
        start_x = event.x
        start_y = event.y

def update_measure(event):
    global end_x, end_y, measuring
    if measuring and start_x is not None:
        end_x = event.x
        end_y = event.y

def end_measure(event):
    global measuring, start_x, start_y, end_x, end_y, orig_w, orig_h, label_w, label_h, measure_message_shown
    if not measuring or start_x is None or end_x is None or orig_w is None or orig_h is None or label_w is None or label_h is None:
        if measuring:
            messagebox.showwarning("Cảnh báo", "Không thể đo: Vui lòng đảm bảo đã chọn hai điểm và khung hình video đang hoạt động.")
        measuring = False
        start_x, start_y, end_x, end_y = None, None, None, None
        measure_message_shown = False
        return

    start_x_orig = int(start_x * orig_w / label_w)
    start_y_orig = int(start_y * orig_h / label_h)
    end_x_orig = int(end_x * orig_w / label_w)
    end_y_orig = int(end_y * orig_h / label_h)

    dist_px = np.sqrt((end_x_orig - start_x_orig)**2 + (end_y_orig - start_y_orig)**2)

    height_input = entry_height_cm.get().strip()
    if height_input:
        try:
            height_cm = float(height_input)
            if height_cm <= 0:
                raise ValueError("Chiều cao phải lớn hơn 0.")
            height_px = orig_h
            scale = height_cm / height_px
            dist_cm = dist_px * scale
            msg = f"Khoảng cách: {dist_cm:.2f} cm ({dist_px:.2f} pixel)"
        except ValueError:
            msg = f"Khoảng cách: {dist_px:.2f} pixel (Chiều cao không hợp lệ: '{height_input}')"
    else:
        msg = f"Khoảng cách: {dist_px:.2f} pixel (Chưa nhập chiều cao)"

    messagebox.showinfo("Số đo", msg)

    if 'dist_cm' in locals():
        field = simpledialog.askstring("Nhập vào trường", "Nhập số này vào trường nào? (vai, hong, do nghien, chieu cao)")
        if field == "vai":
            entry_shoulder_cm.delete(0, tk.END)
            entry_shoulder_cm.insert(0, f"{dist_cm:.2f}")
        elif field == "hong":
            entry_hip_cm.delete(0, tk.END)
            entry_hip_cm.insert(0, f"{dist_cm:.2f}")
        elif field == "chieu cao":
            entry_height_cm.delete(0, tk.END)
            entry_height_cm.insert(0, f"{dist_cm:.2f}")
        elif field == "do nghien":
            entry_tilt_deg.delete(0, tk.END)
            entry_tilt_deg.insert(0, f"{dist_cm:.2f}")

    measuring = False
    start_x, start_y, end_x, end_y = None, None, None, None
    measure_message_shown = False

# =========================
# Cửa sổ So sánh thực tế - nhập tay
# =========================
compare_frame = tk.Frame(compare_window, bg='#2e2e2e')
compare_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

tk.Label(compare_frame, text="Khoảng cách vai (cm):", fg='white', bg='#2e2e2e').grid(row=0, column=0, sticky='w')
entry_shoulder_cm = tk.Entry(compare_frame, width=20)
entry_shoulder_cm.grid(row=0, column=1, padx=5, pady=2)

tk.Label(compare_frame, text="Khoảng cách hông (cm):", fg='white', bg='#2e2e2e').grid(row=1, column=0, sticky='w')
entry_hip_cm = tk.Entry(compare_frame, width=20)
entry_hip_cm.grid(row=1, column=1, padx=5, pady=2)

tk.Label(compare_frame, text="Độ nghiêng vai (°):", fg='white', bg='#2e2e2e').grid(row=2, column=0, sticky='w')
entry_tilt_deg = tk.Entry(compare_frame, width=20)
entry_tilt_deg.grid(row=2, column=1, padx=5, pady=2)

tk.Label(compare_frame, text="Chiều cao (cm):", fg='white', bg='#2e2e2e').grid(row=3, column=0, sticky='w')
entry_height_cm = tk.Entry(compare_frame, width=20)
entry_height_cm.grid(row=3, column=1, padx=5, pady=2)

btn_compare_real = tk.Button(compare_frame, text="📊 So sánh AI – Thực tế",
                            command=lambda: save_ai_vs_real_compare(patient_info, latest_metrics, latest_angle_deg),
                            bg='#9C27B0', fg='white', font=('Arial', 10, 'bold'))
btn_compare_real.grid(row=4, column=0, columnspan=2, pady=10)

btn_ruler = tk.Button(compare_frame, text="📏 Thước đo",
                      command=toggle_measure_mode,
                      bg='#9C27B0', fg='white', font=('Arial', 10, 'bold'))
btn_ruler.grid(row=5, column=0, columnspan=2, pady=10)

# =========================
# Load thông tin bệnh nhân
# =========================
def load_existing_patient_info(*args):
    name = entry_name.get().strip()
    if not name:
        return

    filename = f"{name.replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        if "Thông tin bệnh nhân" in wb.sheetnames:
            info_sheet = wb["Thông tin bệnh nhân"]
            entry_dob.delete(0, tk.END)
            entry_dob.insert(0, info_sheet["B2"].value or "")
            entry_joint.delete(0, tk.END)
            entry_joint.insert(0, info_sheet["B3"].value or "")
            entry_condition.delete(0, tk.END)
            entry_condition.insert(0, info_sheet["B4"].value or "")
    except FileNotFoundError:
        print(f"🆕 Chưa có thông tin cũ cho {name}.")

def load_patient_from_history(*args):
    selected_patient = combo_patient_history.get()
    if selected_patient == "Chọn bệnh nhân..." or selected_patient == "Không có bệnh nhân":
        return

    filename = f"{selected_patient.replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        if "Thông tin bệnh nhân" in wb.sheetnames:
            info_sheet = wb["Thông tin bệnh nhân"]
            entry_name.delete(0, tk.END)
            entry_name.insert(0, info_sheet["B1"].value or "")
            entry_dob.delete(0, tk.END)
            entry_dob.insert(0, info_sheet["B2"].value or "")
            entry_joint.delete(0, tk.END)
            entry_joint.insert(0, info_sheet["B3"].value or "")
            entry_condition.delete(0, tk.END)
            entry_condition.insert(0, info_sheet["B4"].value or "")
    except FileNotFoundError:
        print(f"🆕 Không tìm thấy file cho {selected_patient}.")

entry_name.bind("<FocusOut>", load_existing_patient_info)
entry_name.bind("<Return>", load_existing_patient_info)
combo_patient_history.bind("<<ComboboxSelected>>", load_patient_from_history)

def save_patient_info():
    patient_info["name"] = entry_name.get().strip()
    patient_info["dob"] = entry_dob.get().strip()
    patient_info["joint"] = entry_joint.get().strip()
    patient_info["condition"] = entry_condition.get().strip()

    if not patient_info["name"]:
        messagebox.showerror("Lỗi", "Tên bệnh nhân không được để trống!")
        return False

    filename = f"{patient_info['name'].replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if "Thông tin bệnh nhân" not in wb.sheetnames:
        wb.create_sheet("Thông tin bệnh nhân")
    info_sheet = wb["Thông tin bệnh nhân"]

    info_sheet["A1"] = "Tên bệnh nhân"
    info_sheet["B1"] = patient_info["name"]
    info_sheet["A2"] = "Ngày sinh"
    info_sheet["B2"] = patient_info["dob"]
    info_sheet["A3"] = "Khớp kiểm tra"
    info_sheet["B3"] = patient_info["joint"]
    info_sheet["A4"] = "Tình trạng"
    info_sheet["B4"] = patient_info["condition"]
    info_sheet["A5"] = "Ngày lưu"
    info_sheet["B5"] = datetime.now().strftime("%d/%m/%Y")

    wb.save(filename)
    full_path = os.path.abspath(filename)
    messagebox.showinfo("Lưu thành công", f"Thông tin bệnh nhân đã được lưu vào: {full_path}")
    load_patient_history()
    return True

# =========================
# Biến trạng thái
# =========================
running = True
paused = False
selected_joints = set()
last_person_detected = False
no_person_count = 0

main_frame = tk.Frame(root, bg='#2e2e2e')
main_frame.pack(fill=tk.BOTH, expand=True)

video_frame = tk.Label(video_window, bg='#2e2e2e')
video_frame.pack(fill=tk.BOTH, expand=True)

graph_frame = tk.Frame(main_frame, bg='#2e2e2e')
graph_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

fig, ax = plt.subplots(figsize=(6, 6), facecolor='#2e2e2e')
fig.patch.set_alpha(0.7)
canvas = FigureCanvasTkAgg(fig, master=graph_frame)
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

control_frame = tk.Frame(graph_frame, bg='#2e2e2e')
control_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)

def toggle_pause():
    global paused
    paused = not paused

def close_app():
    global running
    running = False
    root.quit()
    try:
        video_window.destroy()
    except Exception:
        pass
    if cap.isOpened():
        cap.release()
    cv2.destroyAllWindows()

def save_landmark_data(landmark_history, selected_joints, patient_info):
    if not patient_info["name"]:
        messagebox.showerror("Lỗi", "Tên bệnh nhân không được để trống!")
        return False

    date_str = datetime.now().strftime("%Y-%m-%d")
    sheet_name = f"Pose Data - {date_str}"
    graph_filename = f"{patient_info['name'].replace(' ', '_')}_Graph_{date_str}.png"

    filename = f"{patient_info['name'].replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        header = ["Frame"] + [f"Khớp {i} ({ten_khop.get(i, f'Khớp {i}')})" for i in selected_joints]
        ws.append(header)
    else:
        ws = wb[sheet_name]

    if "Thông tin bệnh nhân" not in wb.sheetnames:
        info_sheet = wb.create_sheet("Thông tin bệnh nhân")
        info_sheet["A1"] = "Tên bệnh nhân"
        info_sheet["B1"] = patient_info["name"]
        info_sheet["A2"] = "Ngày sinh"
        info_sheet["B2"] = patient_info["dob"]
        info_sheet["A3"] = "Khớp kiểm tra"
        info_sheet["B3"] = patient_info["joint"]
        info_sheet["A4"] = "Tình trạng"
        info_sheet["B4"] = patient_info["condition"]
        info_sheet["A5"] = "Ngày lưu"
        info_sheet["B5"] = datetime.now().strftime("%d/%m/%Y")

    start_index = ws.max_row
    max_len = max((len(landmark_history[i]) for i in selected_joints), default=0)
    for i in range(max_len):
        row = [start_index + i]
        for j in selected_joints:
            row.append(landmark_history[j][i] if i < len(landmark_history[j]) else "")
        ws.append(row)

    wb.save(filename)
    full_path = os.path.abspath(filename)

    fig.savefig(graph_filename, facecolor=fig.get_facecolor(), edgecolor='none')
    graph_full_path = os.path.abspath(graph_filename)

    messagebox.showinfo("Lưu thành công", f"Dữ liệu theo ngày đã lưu vào sheet '{sheet_name}' trong: {full_path}\nĐồ thị lưu vào: {graph_full_path}")

    load_patient_history()
    return True

def plot_combined_history(patient_name, selected_joints):
    filename = f"{patient_name.replace(' ', '_')}.xlsx"
    if not os.path.exists(filename):
        messagebox.showerror("Lỗi", f"Không tìm thấy file dữ liệu cho {patient_name}")
        return

    wb = openpyxl.load_workbook(filename)
    pose_sheets = [s for s in wb.sheetnames if s.startswith("Pose Data -")]

    if not pose_sheets:
        messagebox.showwarning("Thông báo", "Chưa có dữ liệu Pose nào để tổng hợp.")
        return

    fig2, ax2 = plt.subplots(figsize=(8, 6))
    colors = plt.cm.tab10(np.linspace(0, 1, len(pose_sheets)))

    for idx, sheet_name in enumerate(pose_sheets):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        for j in selected_joints:
            joint_name = f"Khớp {j} ({ten_khop.get(j, f'Khớp {j}')})"
            if joint_name in headers:
                col_idx = headers.index(joint_name)
                frames, values = [], []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    frames.append(row[0])
                    values.append(row[col_idx])
                ax2.plot(frames, values, label=f"{joint_name} - {sheet_name}", color=colors[idx])

    ax2.set_title(f"Tổng hợp dữ liệu khớp - {patient_name}")
    ax2.set_xlabel("Frame")
    ax2.set_ylabel("Tọa độ Y (chuẩn hóa)")
    ax2.legend(fontsize=8)
    ax2.grid(True)

    plt.show()

# =========================
# Bộ nhớ đệm cho smoothing nhiều frame
# =========================
spine_history = deque(maxlen=15)  # Giảm maxlen để phản hồi nhanh hơn

# Biến để lưu previous values cho trường hợp visibility thấp
prev_neck_x = None
prev_neck_y = None
prev_pelvis_x = None
prev_pelvis_y = None
prev_shoulder_angle = 0
prev_hip_angle = 0

# Hằng số alpha cho exponential moving average
alpha = 0.6  # Giá trị alpha cao hơn để cập nhật nhanh hơn

# =========================
# Các hàm phục vụ chụp ảnh + đánh giá cột sống
# =========================
def compute_spine_metrics(spine_x_norm, neck_x_norm, pelvis_x_norm, w_px, shoulder_lm, hip_lm):
    straight_line_x = np.linspace(neck_x_norm * w_px, pelvis_x_norm * w_px, len(spine_x_norm))
    spine_x_px = spine_x_norm * w_px

    deviations = spine_x_px - straight_line_x
    abs_dev = np.abs(deviations)

    avg_abs_dev_px = float(np.mean(abs_dev))
    std_dev_px = float(np.std(deviations))
    max_idx = int(np.argmax(abs_dev))
    max_dev_px = float(abs_dev[max_idx])

    if np.isclose(deviations[max_idx], 0, atol=1e-6):
        direction = "cân bằng"
    else:
        direction = "phải" if deviations[max_idx] > 0 else "trái"

    shoulder_width_px = abs((shoulder_lm[1] - shoulder_lm[0]) * w_px)
    if shoulder_width_px < 1:
        shoulder_width_px = 1.0
    dev_percent_shoulder = 100.0 * avg_abs_dev_px / shoulder_width_px

    if dev_percent_shoulder < 2:
        severity = "nhẹ"
    elif dev_percent_shoulder < 5:
        severity = "vừa"
    elif dev_percent_shoulder < 8:
        severity = "tương đối nặng"
    else:
        severity = "nặng"

    return {
        "avg_abs_dev_px": avg_abs_dev_px,
        "std_dev_px": abs(std_dev_px),
        "max_dev_px": max_dev_px,
        "max_idx": max_idx,
        "direction": direction,
        "severity": severity,
        "dev_percent_shoulder": dev_percent_shoulder,
    }

def draw_spine_overlay(img_bgr, spine_points_xy_px, neck_xy_px, pelvis_xy_px, verdict_text, max_idx=None):
    out = img_bgr.copy()
    cv2.line(out, neck_xy_px, pelvis_xy_px, (255, 255, 0), 2)

    w = out.shape[1]
    thresh1 = 0.02 * w
    thresh2 = 0.05 * w

    for i, (x, y) in enumerate(spine_points_xy_px):
        color = (0, 255, 0)
        if max_idx is not None and i == max_idx:
            color = (0, 0, 255)
        cv2.circle(out, (x, y), 5, color, -1)
        if i < len(spine_points_xy_px) - 1:
            cv2.line(out, spine_points_xy_px[i], spine_points_xy_px[i + 1], (0, 255, 255), 2)

    return out

def snapshot_and_check_spine():
    global current_frame, current_spine_points_x, current_spine_points_y
    global current_neck_x, current_neck_y, current_pelvis_x, current_pelvis_y
    global left_shoulder, right_shoulder, left_hip, right_hip
    global latest_metrics, latest_angle_deg

    if current_frame is None or current_spine_points_x is None or current_neck_x is None:
        messagebox.showerror("Lỗi", "Không có dữ liệu khung hình hoặc cột sống để chụp! Vui lòng đảm bảo camera đang hoạt động.")
        return

    if not patient_info.get("name"):
        if not save_patient_info():
            return

    img_rgb = current_frame.copy()
    h, w, _ = img_rgb.shape

    shoulder_lm = (left_shoulder.x if left_shoulder else 0, right_shoulder.x if right_shoulder else 0)
    hip_lm = (left_hip.x if left_hip else 0, right_hip.x if right_hip else 0)
    spine_history.append((np.array(current_spine_points_x), current_neck_x, current_pelvis_x))
    avg_spine_x = np.mean([s[0] for s in spine_history], axis=0)
    avg_neck_x = np.mean([s[1] for s in spine_history])
    avg_pelvis_x = np.mean([s[2] for s in spine_history])

    metrics = compute_spine_metrics(
        spine_x_norm=avg_spine_x,
        neck_x_norm=avg_neck_x,
        pelvis_x_norm=avg_pelvis_x,
        w_px=w,
        shoulder_lm=shoulder_lm,
        hip_lm=hip_lm,
    )

    z_diff_shoulders = abs(left_shoulder.z - right_shoulder.z) if left_shoulder and right_shoulder else 0
    z_diff_hips = abs(left_hip.z - right_hip.z) if left_hip and right_hip else 0
    rotated = (z_diff_shoulders > 0.1 or z_diff_hips > 0.1)

    shoulder_vec = np.array([right_shoulder.x - left_shoulder.x, right_shoulder.y - left_shoulder.y]) if left_shoulder and right_shoulder else np.array([0.1, 0])
    hip_vec = np.array([right_hip.x - left_hip.x, right_hip.y - left_hip.y]) if left_hip and right_hip else np.array([0.1, 0])
    cosang = np.clip(np.dot(shoulder_vec, hip_vec) / (np.linalg.norm(shoulder_vec) * np.linalg.norm(hip_vec) + 1e-9), -1.0, 1.0)
    angle_deg = float(np.degrees(np.arccos(cosang)))

    latest_metrics = metrics
    latest_angle_deg = angle_deg

    if metrics["dev_percent_shoulder"] < 3 and angle_deg < 5:
        verdict = "✅ Cột sống thẳng, cân bằng"
    else:
        verdict = (f"Lệch {metrics['severity']} về bên {metrics['direction']} | "
                   f"Δtb={metrics['avg_abs_dev_px']:.1f}px (~{metrics['dev_percent_shoulder']:.1f}% bề ngang vai) | "
                   f"Vai–Hông lệch {angle_deg:.1f}°")

    neck_px = (int(current_neck_x * w), int(current_neck_y * h))
    pelvis_px = (int(current_pelvis_x * w), int(current_pelvis_y * h))
    spine_pixels = [(int(x * w), int(y * h)) for x, y in zip(current_spine_points_x, current_spine_points_y)]
    img_bgr = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2BGR)
    annotated = draw_spine_overlay(img_bgr, spine_pixels, neck_px, pelvis_px, verdict, max_idx=metrics['max_idx'])

    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    snapshot_filename = f"snapshot_{patient_info['name'].replace(' ', '_')}_{date_str}.png"
    cv2.imwrite(snapshot_filename, annotated)
    full_path = os.path.abspath(snapshot_filename)

    filename = f"{patient_info['name'].replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    sheet_name = f"Đánh giá cột sống - {datetime.now().strftime('%Y-%m-%d')}"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append([
            "Thời điểm",
            "Ảnh chụp",
            "Trung bình lệch (px)",
            "% so với bề ngang vai",
            "Độ lệch chuẩn (px)",
            "Lệch tối đa (px)",
            "Vị trí max (chỉ số đốt)",
            "Bên lệch",
            "Mức độ",
            "Góc Vai–Hông (°)",
            "Kết luận",
        ])
    else:
        ws = wb[sheet_name]

    ws.append([
        datetime.now().strftime("%H:%M:%S"),
        os.path.abspath(snapshot_filename),
        round(metrics['avg_abs_dev_px'], 2),
        round(metrics['dev_percent_shoulder'], 2),
        round(metrics['std_dev_px'], 2),
        round(metrics['max_dev_px'], 2),
        metrics['max_idx'] + 1,
        metrics['direction'],
        metrics['severity'],
        round(angle_deg, 1),
        verdict,
    ])

    if "thẳng" not in verdict:
        ws.append(["Đánh giá tổng quát:", f"Lệch về bên {metrics['direction']}"])

    wb.save(filename)

    message = (
        f"Ảnh đã lưu: {full_path}\n"
        f"Kết quả: {verdict}\n"
        f"Dữ liệu đã ghi vào '{sheet_name}' của file: {os.path.abspath(filename)}"
    )
    messagebox.showinfo("Kết quả chụp ảnh", message)

def save_ai_vs_real_compare(patient_info, metrics, angle_deg):
    global current_frame, current_spine_points_x, current_neck_x, current_pelvis_x
    global left_shoulder, right_shoulder, left_hip, right_hip

    # Cập nhật patient_info từ các trường nhập liệu
    patient_info["name"] = entry_name.get().strip()
    patient_info["dob"] = entry_dob.get().strip()
    patient_info["joint"] = entry_joint.get().strip()
    patient_info["condition"] = entry_condition.get().strip()

    if not patient_info["name"]:
        messagebox.showerror("Lỗi", "Tên bệnh nhân không được để trống!")
        return False

    # Tự động lưu thông tin bệnh nhân
    saved = save_patient_info()
    if not saved:
        return False

    # Lấy dữ liệu nhập tay
    try:
        shoulder_cm = float(entry_shoulder_cm.get().strip() or 0)
        hip_cm = float(entry_hip_cm.get().strip() or 0)
        tilt_deg_real = float(entry_tilt_deg.get().strip() or 0)
        height_cm = float(entry_height_cm.get().strip() or 0)
    except ValueError:
        messagebox.showerror("Lỗi", "Vui lòng nhập số hợp lệ cho các trường!")
        return False

    if all(x == 0 for x in [shoulder_cm, hip_cm, tilt_deg_real, height_cm]):
        messagebox.showerror("Lỗi", "Vui lòng nhập ít nhất một trường để so sánh!")
        return False

    # Tính toán số liệu cột sống từ khung hình hiện tại hoặc latest_metrics
    if (current_frame is not None and current_spine_points_x is not None and
        current_neck_x is not None and (left_shoulder or right_shoulder) and (left_hip or right_hip)):
        h, w, _ = current_frame.shape
        shoulder_lm = (left_shoulder.x if left_shoulder else right_shoulder.x, right_shoulder.x if right_shoulder else left_shoulder.x)
        hip_lm = (left_hip.x if left_hip else right_hip.x, right_hip.x if right_hip else left_hip.x)
        spine_history.append((np.array(current_spine_points_x), current_neck_x, current_pelvis_x))
        avg_spine_x = np.mean([s[0] for s in spine_history], axis=0)
        avg_neck_x = np.mean([s[1] for s in spine_history])
        avg_pelvis_x = np.mean([s[2] for s in spine_history])

        metrics = compute_spine_metrics(
            spine_x_norm=avg_spine_x,
            neck_x_norm=avg_neck_x,
            pelvis_x_norm=avg_pelvis_x,
            w_px=w,
            shoulder_lm=shoulder_lm,
            hip_lm=hip_lm,
        )

        shoulder_vec = np.array([right_shoulder.x - left_shoulder.x, right_shoulder.y - left_shoulder.y]) if left_shoulder and right_shoulder else np.array([0.1, 0])
        hip_vec = np.array([right_hip.x - left_hip.x, right_hip.y - left_hip.y]) if left_hip and right_hip else np.array([0.1, 0])
        cosang = np.clip(np.dot(shoulder_vec, hip_vec) / (np.linalg.norm(shoulder_vec) * np.linalg.norm(hip_vec) + 1e-9), -1.0, 1.0)
        angle_deg = float(np.degrees(np.arccos(cosang)))
    elif metrics is None or angle_deg is None:
        messagebox.showerror("Lỗi", "Không có dữ liệu cột sống từ AI! Vui lòng chụp ảnh cột sống trước.")
        return False

    # Chuyển đổi số liệu AI sang cm nếu có chiều cao
    scale = height_cm / h if height_cm > 0 and 'h' in locals() else None
    ai_metrics = {
        "shoulder_cm": metrics["avg_abs_dev_px"] * scale if scale else "-",
        "hip_cm": metrics["max_dev_px"] * scale if scale else "-",
        "tilt_deg": round(angle_deg, 2) if angle_deg is not None else "-",
        "height_cm": height_cm if height_cm > 0 else "-"
    }

    # Tính sai số phần trăm
    percent_errors = {}
    if shoulder_cm > 0 and ai_metrics["shoulder_cm"] != "-":
        percent_errors["shoulder"] = abs(ai_metrics["shoulder_cm"] - shoulder_cm) / shoulder_cm * 100
    if hip_cm > 0 and ai_metrics["hip_cm"] != "-":
        percent_errors["hip"] = abs(ai_metrics["hip_cm"] - hip_cm) / hip_cm * 100
    if tilt_deg_real > 0 and ai_metrics["tilt_deg"] != "-":
        percent_errors["tilt"] = abs(ai_metrics["tilt_deg"] - tilt_deg_real) / tilt_deg_real * 100
    if height_cm > 0 and ai_metrics["height_cm"] != "-":
        percent_errors["height"] = abs(ai_metrics["height_cm"] - height_cm) / height_cm * 100

    # Tạo kết luận
    if percent_errors:
        avg_percent_error = sum(percent_errors.values()) / len(percent_errors)
        conclusion = f"Sai số trung bình: {avg_percent_error:.2f}% so với thực tế"
    else:
        conclusion = "Không có sai số do thiếu dữ liệu so sánh."

    # Lưu vào Excel
    filename = f"{patient_info['name'].replace(' ', '_')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        messagebox.showerror("Lỗi", f"Không tìm thấy file cho bệnh nhân {patient_info['name']}")
        return False

    sheet_name = f"So sánh AI - Thực tế"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append([
            "Thời điểm",
            "Vai (cm) [AI]",
            "Vai (cm) [Thực tế]",
            "Sai số vai (%)",
            "Hông (cm) [AI]",
            "Hông (cm) [Thực tế]",
            "Sai số hông (%)",
            "Độ nghiêng vai (°) [AI]",
            "Độ nghiêng vai (°) [Thực tế]",
            "Sai số độ nghiêng (%)",
            "Chiều cao (cm) [AI]",
            "Chiều cao (cm) [Thực tế]",
            "Sai số chiều cao (%)",
            "Kết luận"
        ])
    else:
        ws = wb[sheet_name]

    ws.append([
        datetime.now().strftime("%H:%M:%S"),
        round(ai_metrics["shoulder_cm"], 2) if ai_metrics["shoulder_cm"] != "-" else "-",
        shoulder_cm if shoulder_cm > 0 else "-",
        round(percent_errors["shoulder"], 2) if "shoulder" in percent_errors else "-",
        round(ai_metrics["hip_cm"], 2) if ai_metrics["hip_cm"] != "-" else "-",
        hip_cm if hip_cm > 0 else "-",
        round(percent_errors["hip"], 2) if "hip" in percent_errors else "-",
        ai_metrics["tilt_deg"],
        tilt_deg_real if tilt_deg_real > 0 else "-",
        round(percent_errors["tilt"], 2) if "tilt" in percent_errors else "-",
        ai_metrics["height_cm"],
        height_cm if height_cm > 0 else "-",
        round(percent_errors["height"], 2) if "height" in percent_errors else "-",
        conclusion
    ])

    wb.save(filename)
    messagebox.showinfo("Thành công", f"So sánh AI – Thực tế đã lưu vào sheet '{sheet_name}' trong file: {os.path.abspath(filename)}\n\n{conclusion}")
    return True

# =========================
# Các nút điều khiển
# =========================
btn_toggle = tk.Button(control_frame, text="⏯ Tạm dừng / Chạy", command=toggle_pause,
                      bg='#3e3e3e', fg='white', font=('Arial', 10, 'bold'))
btn_toggle.pack(side=tk.LEFT, padx=10)

btn_exit = tk.Button(control_frame, text="❌ Thoát", command=close_app,
                     bg='#6e2e2e', fg='white', font=('Arial', 10, 'bold'))
btn_exit.pack(side=tk.LEFT, padx=10)

btn_save = tk.Button(control_frame, text="📀 Lưu dữ liệu",
                     command=lambda: [save_patient_info(), save_landmark_data(landmark_history, selected_joints, patient_info)],
                     bg='#2196F3', fg='white', font=('Arial', 10, 'bold'))
btn_save.pack(side=tk.LEFT, padx=10)

btn_combine = tk.Button(control_frame, text="📊 Tổng hợp nhiều ngày",
                        command=lambda: plot_combined_history(patient_info["name"], selected_joints),
                        bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'))
btn_combine.pack(side=tk.LEFT, padx=10)

btn_snapshot = tk.Button(control_frame, text="📸 Chụp + Đánh giá + Ghi Excel", command=snapshot_and_check_spine,
                         bg='#FF5722', fg='white', font=('Arial', 10, 'bold'))
btn_snapshot.pack(side=tk.LEFT, padx=10)

# =========================
# Khởi tạo dữ liệu
# =========================
num_spine_segments = 17
total_joints = 33 + num_spine_segments
landmark_history = [[] for _ in range(total_joints)]
max_points = 100
colors = plt.cm.viridis(np.linspace(0, 1, total_joints))

ten_khop = {
    0: "Mũi", 1: "Mắt trái trong", 2: "Mắt trái", 3: "Mắt trái ngoài",
    4: "Mắt phải trong", 5: "Mắt phải", 6: "Mắt phải ngoài",
    7: "Tai trái", 8: "Tai phải", 9: "Miệng trái", 10: "Miệng phải",
    11: "Vai trái", 12: "Vai phải", 13: "Khuỷu tay trái",
    14: "Khuỷu tay phải", 15: "Cổ tay trái", 16: "Cổ tay phải",
    17: "Ngón út tay trái", 18: "Ngón út tay phải",
    19: "Ngón trỏ tay trái", 20: "Ngón trỏ tay phải",
    21: "Ngón cái tay trái", 22: "Ngón cái tay phải",
    23: "Hông trái", 24: "Hông phải", 25: "Gối trái",
    26: "Gối phải", 27: "Cổ chân trái", 28: "Cổ chân phải",
    29: "Gót trái", 30: "Gót phải", 31: "Ngón chân trái", 32: "Ngón chân phải"
}
for i in range(num_spine_segments):
    if i < 12:
        ten_khop[33 + i] = f"Đốt ngực T{i+1}"
    else:
        ten_khop[33 + i] = f"Đốt thắt lưng L{i-11}"

joint_groups = {
    "Đầu": [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    "Tay trái": [11, 13, 15, 17, 19, 21],
    "Tay phải": [12, 14, 16, 18, 20, 22],
    "Hông": [23, 24],
    "Chân trái": [25, 27, 29, 31],
    "Chân phải": [26, 28, 30, 32],
    "Cột sống": list(range(33, 33 + num_spine_segments))
}

lines = []
for i in range(total_joints):
    (line,) = ax.plot([], [], label=ten_khop.get(i, f"Khớp {i}"), color=colors[i], linewidth=1.5)
    lines.append(line)

ax.set_facecolor('#3e3e3e')
for spine in ax.spines.values():
    spine.set_color('#5e5e5e')
    spine.set_linewidth(2)

ax.set_xlim(0, max_points)
ax.set_ylim(0, 1)
ax.set_title("ĐỒ THỊ THEO DÕI KHỚP")
ax.set_xlabel("Frame")
ax.set_ylabel("Tọa độ Y")

checkbox_frame = tk.Frame(graph_frame, bg='#2e2e2e')
checkbox_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5)

tk.Label(checkbox_frame, text="Chọn nhóm khớp:", fg='white', bg='#2e2e2e').pack(anchor='w', padx=5, pady=2)
combo_joint_group = ttk.Combobox(checkbox_frame, values=list(joint_groups.keys()), state="readonly")
combo_joint_group.pack(fill='x', padx=5, pady=2)
combo_joint_group.set("Chọn nhóm...")

canvas_scroll = tk.Canvas(checkbox_frame, bg='#2e2e2e', highlightthickness=0)
scrollbar = ttk.Scrollbar(checkbox_frame, orient="vertical", command=canvas_scroll.yview)
scrollable_frame = tk.Frame(canvas_scroll, bg='#2e2e2e')

scrollable_frame.bind("<Configure>", lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
canvas_scroll.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas_scroll.configure(yscrollcommand=scrollbar.set)
canvas_scroll.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

checkbox_vars = []
checkboxes = []

def update_selected_joints(index, var):
    if var.get():
        selected_joints.add(index)
    else:
        selected_joints.discard(index)

for i in range(total_joints):
    var = tk.IntVar()
    cb = tk.Checkbutton(scrollable_frame, text=ten_khop.get(i, f"Khớp {i}"), variable=var,
                        command=lambda i=i, v=var: update_selected_joints(i, v),
                        bg='#2e2e2e', fg='white', selectcolor='#444444')
    checkbox_vars.append(var)
    checkboxes.append(cb)

def update_checkboxes(*args):
    for cb in checkboxes:
        cb.pack_forget()
    selected_group = combo_joint_group.get()
    if selected_group in joint_groups:
        for i in joint_groups[selected_group]:
            checkboxes[i].pack(fill='x', padx=2, pady=1)

combo_joint_group.bind("<<ComboboxSelected>>", update_checkboxes)

# =========================
# Khởi tạo camera
# =========================
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    messagebox.showerror("Lỗi", "Không thể mở camera! Vui lòng kiểm tra kết nối camera hoặc quyền truy cập.")
    root.destroy()
    raise SystemExit

# =========================
# Phím tắt
# =========================
def on_key_press(event):
    global running
    ch = event.char.lower() if event.char else ''
    if ch == 'q':
        running = False
        root.quit()
        try:
            video_window.destroy()
        except Exception:
            pass
    elif ch == 'c':
        snapshot_and_check_spine()

root.bind('<Key>', on_key_press)
video_window.bind('<Key>', on_key_press)

video_frame.bind("<ButtonPress-1>", start_measure)
video_frame.bind("<Motion>", update_measure)
video_frame.bind("<ButtonRelease-1>", end_measure)

# =========================
# Biến dùng trong vòng lặp
# =========================
current_frame = None
current_spine_points_x = None
current_spine_points_y = None
current_neck_x = None
current_neck_y = None
current_pelvis_x = None
current_pelvis_y = None
left_shoulder = None
right_shoulder = None
left_hip = None
right_hip = None

# =========================
# Vòng lặp cập nhật khung hình
# =========================
def update_frame():
    global running, paused, last_person_detected, no_person_count
    global current_frame, current_spine_points_x, current_spine_points_y
    global current_neck_x, current_neck_y, current_pelvis_x, current_pelvis_y
    global left_shoulder, right_shoulder, left_hip, right_hip
    global orig_w, orig_h, label_w, label_h
    global prev_neck_x, prev_neck_y, prev_pelvis_x, prev_pelvis_y
    global prev_shoulder_angle, prev_hip_angle

    if not running:
        if cap.isOpened():
            cap.release()
        cv2.destroyAllWindows()
        try:
            root.destroy()
            video_window.destroy()
        except Exception:
            pass
        return

    if not paused:
        ret, frame = cap.read()
        if not ret:
            messagebox.showwarning("Cảnh báo", "Không thể đọc khung hình từ camera! Chương trình sẽ tiếp tục chạy nhưng không hiển thị video.")
            root.after(10, update_frame)
            return

        frame = cv2.flip(frame, 1)
        image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        results = pose.process(image)

        if results.pose_landmarks:
            last_person_detected = True
            no_person_count = 0
            mp_drawing.draw_landmarks(
                image, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                mp_drawing.DrawingSpec(color=(245, 117, 66), thickness=2, circle_radius=2),
                mp_drawing.DrawingSpec(color=(245, 66, 230), thickness=2, circle_radius=2)
            )
            landmarks = results.pose_landmarks.landmark
            for idx, landmark in enumerate(landmarks):
                landmark_history[idx].append(landmark.y)
                if len(landmark_history[idx]) > max_points:
                    landmark_history[idx].pop(0)

            # Check visibility for shoulders and hips
            visibility_threshold = 0.5
            left_shoulder = landmarks[11] if landmarks[11].visibility > visibility_threshold else None
            right_shoulder = landmarks[12] if landmarks[12].visibility > visibility_threshold else None
            left_hip = landmarks[23] if landmarks[23].visibility > visibility_threshold else None
            right_hip = landmarks[24] if landmarks[24].visibility > visibility_threshold else None

            # Calculate neck and pelvis using visible landmarks only
            shoulder_xs = []
            shoulder_ys = []
            shoulder_zs = []
            if left_shoulder:
                shoulder_xs.append(left_shoulder.x)
                shoulder_ys.append(left_shoulder.y)
                shoulder_zs.append(left_shoulder.z)
            if right_shoulder:
                shoulder_xs.append(right_shoulder.x)
                shoulder_ys.append(right_shoulder.y)
                shoulder_zs.append(right_shoulder.z)

            hip_xs = []
            hip_ys = []
            hip_zs = []
            if left_hip:
                hip_xs.append(left_hip.x)
                hip_ys.append(left_hip.y)
                hip_zs.append(left_hip.z)
            if right_hip:
                hip_xs.append(right_hip.x)
                hip_ys.append(right_hip.y)
                hip_zs.append(right_hip.z)

            if not shoulder_xs or not hip_xs:
                # Use previous if available
                if prev_neck_x is not None and prev_pelvis_x is not None:
                    current_neck_x = prev_neck_x
                    current_neck_y = prev_neck_y
                    current_pelvis_x = prev_pelvis_x
                    current_pelvis_y = prev_pelvis_y
                else:
                    # Skip spine calculation
                    current_spine_points_x = None
                    current_spine_points_y = None
            else:
                new_neck_x = np.mean(shoulder_xs)
                new_neck_y = np.mean(shoulder_ys)
                new_pelvis_x = np.mean(hip_xs)
                new_pelvis_y = np.mean(hip_ys)

                if prev_neck_x is None:
                    current_neck_x = new_neck_x
                    current_neck_y = new_neck_y
                    current_pelvis_x = new_pelvis_x
                    current_pelvis_y = new_pelvis_y
                else:
                    current_neck_x = alpha * new_neck_x + (1 - alpha) * prev_neck_x
                    current_neck_y = alpha * new_neck_y + (1 - alpha) * prev_neck_y
                    current_pelvis_x = alpha * new_pelvis_x + (1 - alpha) * prev_pelvis_x
                    current_pelvis_y = alpha * new_pelvis_y + (1 - alpha) * prev_pelvis_y

                prev_neck_x = current_neck_x
                prev_neck_y = current_neck_y
                prev_pelvis_x = current_pelvis_x
                prev_pelvis_y = current_pelvis_y

                # Calculate z diffs
                z_diff_shoulders = np.std(shoulder_zs) * 2 if len(shoulder_zs) > 1 else 0
                z_diff_hips = np.std(hip_zs) * 2 if len(hip_zs) > 1 else 0

                # Angles only if both visible, else use previous
                shoulder_angle = prev_shoulder_angle
                if left_shoulder and right_shoulder:
                    shoulder_angle = np.arctan2(right_shoulder.y - left_shoulder.y, right_shoulder.x - left_shoulder.x)
                    prev_shoulder_angle = shoulder_angle

                hip_angle = prev_hip_angle
                if left_hip and right_hip:
                    hip_angle = np.arctan2(right_hip.y - left_hip.y, right_hip.x - left_hip.x)
                    prev_hip_angle = hip_angle

                avg_angle = (shoulder_angle + hip_angle) / 2

                # Detect rotated and adjust offset
                rotated = (z_diff_shoulders > 0.1 or z_diff_hips > 0.1)
                offset_factor = 0.5 if rotated else 1.5  # Tăng offset khi không xoay để cong hơn

                h, w, _ = image.shape
                shoulder_diff_px = int(abs(np.mean(shoulder_ys) * h - np.mean(shoulder_ys) * h))  # Simplified, adjust if needed
                lx, ly = int(np.mean(shoulder_xs) * w), int(np.mean(shoulder_ys) * h)
                rx, ry = int(np.mean(shoulder_xs) * w), int(np.mean(shoulder_ys) * h)  # Adjust for single
                color = (0, 255, 0) if shoulder_diff_px < 3 else (0, 0, 255)
                cv2.line(image, (lx, ly), (rx, ry), color, 3)

                # Calculate spine with adjusted offset
                mid_x = (current_neck_x + current_pelvis_x) / 2
                mid_y = (current_neck_y + current_pelvis_y) / 2
                offset_x = float(np.sin(avg_angle) * 0.15 * offset_factor)
                offset_x = float(np.clip(offset_x, -0.075, 0.075))
                mid_control_x = mid_x + offset_x
                mid_control_y = mid_y + abs(np.sin(avg_angle)) * 0.05 * offset_factor  # Thay đổi dấu để điều chỉnh y, tăng cong

                control_points_x = [current_neck_x, mid_control_x, current_pelvis_x]
                control_points_y = [current_neck_y, mid_control_y, current_pelvis_y]

                tck, _ = splprep([control_points_x, control_points_y], s=0.1, k=2)  # Tăng s để smooth hơn
                u_fine = np.linspace(0, 1, num_spine_segments + 2)
                current_spine_points_x, current_spine_points_y = splev(u_fine, tck)
                current_spine_points_x = current_spine_points_x[1:-1]
                current_spine_points_y = current_spine_points_y[1:-1]

                min_x = min(shoulder_xs + hip_xs) - 0.1
                max_x = max(shoulder_xs + hip_xs) + 0.1
                current_spine_points_x = np.clip(current_spine_points_x, min_x, max_x)

                spine_pixels = []
                for idx, (x, y) in enumerate(zip(current_spine_points_x, current_spine_points_y)):
                    landmark_history[33 + idx].append(y)
                    if len(landmark_history[33 + idx]) > max_points:
                        landmark_history[33 + idx].pop(0)
                    cx, cy = int(x * w), int(y * h)
                    cv2.circle(image, (cx, cy), 5, (0, 255, 0), -1)
                    spine_pixels.append((cx, cy))

                for i in range(len(spine_pixels) - 1):
                    cv2.line(image, spine_pixels[i], spine_pixels[i + 1], (0, 255, 255), 2)

            current_frame = image.copy()

            for i in selected_joints:
                x_vals = list(range(len(landmark_history[i])))
                y_vals = landmark_history[i]
                lines[i].set_data(x_vals, y_vals)
            canvas.draw()
        else:
            no_person_count += 1
            if no_person_count > 30:
                if last_person_detected:
                    messagebox.showwarning("Cảnh báo", "Không phát hiện được người! Vui lòng điều chỉnh tư thế để camera nhận diện vai và hông.")
                    if any(len(landmark_history[i]) > 0 for i in selected_joints):
                        saved = save_patient_info()
                        if saved:
                            saved = save_landmark_data(landmark_history, selected_joints, patient_info)
                        if saved:
                            for i in range(total_joints):
                                landmark_history[i].clear()
                            selected_joints.clear()
                            for var in checkbox_vars:
                                var.set(0)
                            combo_joint_group.set("Chọn nhóm...")
                            update_checkboxes()
                last_person_detected = False
                no_person_count = 0

        orig_h, orig_w, _ = image.shape

        image_copy = image.copy()
        if measuring and start_x is not None and end_x is not None:
            start_x_orig = int(start_x * orig_w / label_w)
            start_y_orig = int(start_y * orig_h / label_h)
            end_x_orig = int(end_x * orig_w / label_w)
            end_y_orig = int(end_y * orig_h / label_h)

            cv2.line(image_copy, (start_x_orig, start_y_orig), (end_x_orig, end_y_orig), (0, 255, 255), 3)
            cv2.circle(image_copy, (start_x_orig, start_y_orig), 5, (0, 255, 0), -1)
            cv2.circle(image_copy, (end_x_orig, end_y_orig), 5, (0, 255, 0), -1)

            mid_x = (start_x_orig + end_x_orig) // 2
            mid_y = (start_y_orig + end_y_orig) // 2
            dist = np.sqrt((end_x_orig - start_x_orig)**2 + (end_y_orig - start_y_orig)**2)

            text = f"{dist:.2f} px"
            text_size, _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)
            text_w, text_h = text_size
            text_bg = (mid_x - text_w // 2 - 5, mid_y - text_h - 15, text_w + 10, text_h + 10)
            cv2.rectangle(image_copy, (text_bg[0], text_bg[1]),
                          (text_bg[0] + text_bg[2], text_bg[1] + text_bg[3]),
                          (50, 50, 50, 200), -1)
            cv2.putText(image_copy, text, (mid_x - text_w // 2, mid_y - 5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
        image = image_copy

        screen_height = video_window.winfo_screenheight()
        desired_height = int(screen_height * 0.8)
        img = Image.fromarray(image)
        w_img, h_img = img.size
        aspect_ratio = w_img / h_img
        label_w = int(desired_height * aspect_ratio)
        label_h = desired_height
        img = img.resize((label_w, label_h))

        imgtk = ImageTk.PhotoImage(image=img)
        video_frame.imgtk = imgtk
        video_frame.configure(image=imgtk)

    root.after(10, update_frame)

video_window.protocol("WM_DELETE_WINDOW", close_app)
update_frame()
root.mainloop()

if cap.isOpened():
    cap.release()
cv2.destroyAllWindows()